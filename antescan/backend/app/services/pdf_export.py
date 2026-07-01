"""Report exports — real PDF, CSV, XLSX file generation."""
import os, csv
from io import BytesIO
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
)


GHANA_RED = colors.HexColor('#CE1126')
GHANA_GOLD = colors.HexColor('#FCD116')
GHANA_GREEN = colors.HexColor('#006B3F')


def _styles():
    s = getSampleStyleSheet()
    s.add(ParagraphStyle('GhanaTitle', parent=s['Title'], textColor=GHANA_RED, fontSize=20,
                          spaceAfter=10, fontName='Helvetica-Bold'))
    s.add(ParagraphStyle('GhanaSub', parent=s['Normal'], textColor=colors.grey, fontSize=9,
                          spaceAfter=8, fontName='Helvetica-Oblique'))
    s.add(ParagraphStyle('SectionH', parent=s['Heading2'], textColor=GHANA_GREEN, fontSize=13,
                          spaceBefore=12, spaceAfter=6))
    return s


def patient_report_pdf(patient, screenings, referrals, lab_results, out_path: str):
    """Generate a clinical patient summary PDF."""
    s = _styles()
    doc = SimpleDocTemplate(out_path, pagesize=A4,
                              topMargin=1.5*cm, bottomMargin=1.5*cm,
                              leftMargin=1.5*cm, rightMargin=1.5*cm)
    story = []

    # Header
    story.append(Paragraph('AnteScan — Patient Clinical Summary', s['GhanaTitle']))
    story.append(Paragraph(f"Generated {datetime.utcnow().strftime('%d %B %Y %H:%M')} UTC", s['GhanaSub']))
    story.append(Spacer(1, 12))

    # Patient block
    info = [
        ['Name', patient.full_name],
        ['Age', f'{patient.age}'],
        ['Sex', patient.sex or '—'],
        ['Phone', patient.phone or '—'],
        ['Village', patient.village or '—'],
        ['Module', patient.primary_module],
    ]
    tbl = Table(info, colWidths=[4*cm, 12*cm])
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F5F5F5')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#666666')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#DDDDDD')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 14))

    # Screenings table
    story.append(Paragraph('Screening History', s['SectionH']))
    if screenings:
        rows = [['Date', 'Module', 'Risk Score', 'Risk Level', 'Model']]
        for sc in screenings:
            rows.append([
                sc.created_at.strftime('%Y-%m-%d %H:%M') if sc.created_at else '—',
                sc.module,
                str(sc.risk_score),
                sc.risk_level,
                sc.model_version or 'rules',
            ])
        t = Table(rows, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), GHANA_GOLD),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#DDDDDD')),
            ('PADDING', (0, 0), (-1, -1), 5),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        story.append(t)
    else:
        story.append(Paragraph('No screenings recorded.', s['Normal']))
    story.append(Spacer(1, 14))

    # Referrals
    story.append(Paragraph('Referrals', s['SectionH']))
    if referrals:
        rows = [['Date', 'Facility', 'Urgency', 'Status', 'Days Open']]
        for r in referrals:
            rows.append([
                r.created_at.strftime('%Y-%m-%d') if r.created_at else '—',
                r.facility_name, r.urgency, r.status, str(r.days_open),
            ])
        t = Table(rows, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), GHANA_RED),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#DDDDDD')),
            ('PADDING', (0, 0), (-1, -1), 5),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        story.append(t)
    else:
        story.append(Paragraph('No referrals.', s['Normal']))
    story.append(Spacer(1, 14))

    # Lab results
    story.append(Paragraph('Lab Results', s['SectionH']))
    if lab_results:
        rows = [['Date', 'Test', 'Result', 'Facility']]
        for lr in lab_results:
            rows.append([
                lr.performed_at.isoformat() if lr.performed_at else '—',
                lr.test_type, lr.result, lr.facility or '—',
            ])
        t = Table(rows, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), GHANA_GREEN),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#DDDDDD')),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        story.append(t)
    else:
        story.append(Paragraph('No lab results.', s['Normal']))

    story.append(Spacer(1, 24))
    story.append(Paragraph(
        '<i>Ghana Health Service · CHPS Programme · Generated by AnteScan</i>',
        s['GhanaSub']))

    doc.build(story)
    return out_path


def screenings_csv(screenings, out_path: str):
    with open(out_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['id', 'created_at', 'patient', 'chw', 'module',
                    'risk_score', 'risk_level', 'bp_systolic', 'bp_diastolic',
                    'muac_mm', 'nutri_class', 'model_version'])
        for s in screenings:
            w.writerow([
                s.id, s.created_at.isoformat() if s.created_at else '',
                s.patient.full_name if s.patient else '',
                s.chw.name if s.chw else '',
                s.module, s.risk_score, s.risk_level,
                s.bp_systolic, s.bp_diastolic, s.muac_mm, s.nutri_class,
                s.model_version or 'rules',
            ])
    return out_path


def referrals_csv(referrals, out_path: str):
    with open(out_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['id', 'date', 'patient', 'module', 'urgency', 'facility',
                    'status', 'sms_status', 'days_open', 'elder_notified'])
        for r in referrals:
            w.writerow([
                r.id, r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else '',
                r.patient.full_name if r.patient else '',
                r.module, r.urgency, r.facility_name,
                r.status, r.sms_status, r.days_open, r.elder_notified,
            ])
    return out_path


def chws_xlsx(chws, out_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = 'CHWs'
    headers = ['CHW ID', 'Name', 'Compound', 'District', 'Region',
                'Badge', 'Total Screenings', 'Week Screenings', 'Status', 'Last Active']
    ws.append(headers)
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='FCD116')
        cell.alignment = Alignment(horizontal='center')

    for c in chws:
        ws.append([
            c.chw_id, c.name, c.compound_obj.name if c.compound_obj else '',
            c.district, c.region, c.badge,
            c.total_screenings, c.week_screenings, c.status,
            c.last_active.isoformat() if c.last_active else '',
        ])

    # Auto-size columns
    for col in ws.columns:
        length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(length + 2, 30)

    wb.save(out_path)
    return out_path
